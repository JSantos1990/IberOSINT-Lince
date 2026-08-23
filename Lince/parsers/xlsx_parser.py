from openpyxl import load_workbook


class XLSXParser:

    @staticmethod
    def extract_text(filepath):

        workbook = load_workbook(filepath, data_only=True)

        lines = []

        for sheet in workbook.worksheets:

            lines.append(f"=== Hoja: {sheet.title} ===")

            for row in sheet.iter_rows(values_only=True):

                values = []

                for cell in row:

                    if cell is None:
                        values.append("")
                    else:
                        values.append(str(cell))

                if any(values):
                    lines.append(" | ".join(values))

            lines.append("")

        return "\n".join(lines)